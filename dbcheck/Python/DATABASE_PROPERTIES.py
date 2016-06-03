import collections
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, colors


class DBINFO:
    # 普通字典不进行排序需要，使用OrderedDict保证报表信息顺序一致
    DBDICT = collections.OrderedDict()
    DBDICT['ZWDB'] = {'NAME': 'ZWDB', 'IP': '10.220.83.21', 'VERSION': '11.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['CRMDB'] = {'NAME': 'CRMDB', 'IP': '10.220.50.231', 'VERSION': '10.2.0.5', 'ISCLUSTER': 'YES'}
    DBDICT['JFDB'] = {'NAME': 'JFDB', 'IP': '10.220.83.11', 'VERSION': '11.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['CBOSSDB'] = {'NAME': 'CBOSSDB', 'IP': '10.220.50.211', 'VERSION': '10.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['KFDB'] = {'NAME': 'KFDB', 'IP': '10.220.50.200', 'VERSION': '10.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['NGKTDB'] = {'NAME': 'NGKTDB', 'IP': '10.220.50.36', 'VERSION': '10.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['PBDB'] = {'NAME': 'PBDB', 'IP': '10.220.50.241', 'VERSION': '10.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['JSDB'] = {'NAME': 'JSDB', 'IP': '10.220.83.71', 'VERSION': '11.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['ZWCXDB'] = {'NAME': 'ZWCXDB', 'IP': '10.220.83.41', 'VERSION': '11.2.0.4', 'ISCLUSTER': 'YES'}
    DBDICT['JFCXDB'] = {'NAME': 'JFCXDB', 'IP': '10.220.83.51', 'VERSION': '11.2.0.4', 'ISCLUSTER': 'YES'}

    def __init__(self):
        print(self.DBDICT)


class REPORTINFO:
    TITLE = {
        'REPORTTITLE': '${DBNAME} 数据库 ${LASTDAY} 日常健康检查报告',
        'INSTANCE_STATUS': '实例运行状态',
        'INSTANCE_LOAD': '实例运行负载 (DB time)',
        'TABLESPACE_USAGE': '表空间使用状况',
        'FILESYSTEM_USAGE': '文件系统使用状况',
        'CLUSTER_STATUS': '集群运行状态',
        'ASM_USAGE': 'ASM磁盘组使用状况',
        'ALTER_INFO': '告警日志分析'
    }

    def __init__(self):
        pass


class REPORTSTYLE:
    REPORTTITLE = {}
    INSTANCE_STATUS = {}
    INSTANCE_LOAD = {}
    TABLESPACE_USAGE = {}
    FILESYSTEM_USAGE = {}
    CLUSTER_STATUS = {}
    ASM_USAGE = {}
    ALTER_INFO = {}
    ROWTITLE = {}

    def __init__(self):
        pass

    # 定义表格标题字体
    def TITLEFONT(titleFlag):
        if titleFlag == 'SHEETTITLE':
            return Font(color=colors.BLACK, sz=20, bold=True)
        elif titleFlag == 'AREATITLE':
            return Font(color=colors.BLACK, sz=12, bold=True)
        elif titleFlag == 'TABLETITLE':
            return Font(color=colors.BLACK, sz=11)

    # 定义表格标题对其方式
    def TITLEALIGN(titleFlag=None):
        if titleFlag == 'SHEETTITLE':
            return Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
        elif titleFlag == 'AREATITLE':
            return Alignment(horizontal='left',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
        else:
            return Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)

    # 定义表格背景颜色
    def TITLEFILL(titleFlag):
        if titleFlag == 'SHEETTITLE':
            return PatternFill(patternType='solid', fgColor='EEE9E9')
        elif titleFlag == 'AREATITLE':
            return PatternFill(patternType='solid', fgColor='87CEEB')
        elif titleFlag == 'TABLETITLE':
            return PatternFill(patternType='solid', fgColor='CDC8B1')
        elif titleFlag == 'ALERTTITLE':
            return PatternFill(patternType='solid', fgColor='FF0000')

    # 定义表格区域合并信息
    def TITLEMERGE(worksheet, rownum, colnum, isTitle):
        if isTitle:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        else:
            worksheet.merge_cells(start_row=rownum, start_column=1, end_row=rownum, end_column=10)

    # 定义表格边框样式
    def TABLEBORDER(titleFlag='TABLEBORDER'):
        if titleFlag is None:
            return Border(left=Side(border_style='medium',
                                    color='FF000000'),
                          right=Side(border_style='medium',
                                     color='FF000000'),
                          top=Side(border_style='medium',
                                   color='FF000000'),
                          bottom=Side(border_style='medium',
                                      color='FF000000'),
                          diagonal=Side(border_style='medium',
                                        color='FF000000'),
                          outline=Side(border_style='medium',
                                       color='FF000000'),
                          vertical=Side(border_style='medium',
                                        color='FF000000'),
                          horizontal=Side(border_style='medium',
                                          color='FF000000'),
                          start=Side(border_style='medium',
                                     color='FF000000')
                          )
        else:
            return Border(left=Side(border_style='thin',
                                    color='FF000000'),
                          right=Side(border_style='thin',
                                     color='FF000000'),
                          top=Side(border_style='thin',
                                   color='FF000000'),
                          bottom=Side(border_style='thin',
                                      color='FF000000'),
                          diagonal=Side(border_style='thin',
                                        color='FF000000'),
                          outline=Side(border_style='thin',
                                       color='FF000000'),
                          vertical=Side(border_style='thin',
                                        color='FF000000'),
                          horizontal=Side(border_style='thin',
                                          color='FF000000'),
                          start=Side(border_style='thin',
                                     color='FF000000')
                          )
