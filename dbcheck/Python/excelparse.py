# -*- coding:utf-8 -*-

import os
import dbchart
from openpyxl import load_workbook
from datetime import datetime, timedelta
from DATABASE_PROPERTIES import REPORTINFO, REPORTSTYLE

TODAY = datetime.now()
LASTDAY = (TODAY - timedelta(days=1)).strftime('%Y%m%d')

DB_NAME_TEMPLATE = '${DBNAME}'
LASTDAY_TEMPLATE = '${LASTDAY}'


# 判断文件是否存在
def FILEEXIST(resultday=LASTDAY, dbname='ZWDB', resultname=None):
    filename = './' + resultday + '/' + dbname + '/' + resultname + '.txt'
    if os.path.isfile(filename):
        return True
    else:
        return False


# SQL结果处理程序
def READSQLRESULT(resultday=LASTDAY, dbname='ZWDB', resultname=None):
    filename = './' + resultday + '/' + dbname + '/' + resultname + '.txt'

    if FILEEXIST(resultday, dbname, resultname):
        return open(filename, 'r')
    else:
        return None


def excelReport(dbname='ZWDB', dbcheckws=None, loggerInfo=None):
    # 通用配置信息
    DB_NAME = dbname

    # 表格模版文件名
    loggerInfo.debug("begin read excel template")
    DBBOOK = load_workbook('DBCHECK_TEMPLATE.xlsx')
    loggerInfo.debug("read excel template sucess")

    # 加载模版SHEET信息
    DBSHEET = DBBOOK[DB_NAME_TEMPLATE]

    # 生成报表文件标题名称
    DBTITLE = DBSHEET.title
    DBTITLE = DBTITLE.replace(DB_NAME_TEMPLATE, DB_NAME)
    loggerInfo.debug("set excel sheet name = %s " % DBTITLE)
    dbcheckws.title = DBTITLE

    # 定义表格开始行、列位置信息
    rowcount = 1
    newrownum = 1

    # 定义表格模块标识信息
    tabsplit = ''

    loggerInfo.debug("Begin read excel template rows")
    # 遍历模版文件行信息
    for DBWROW in DBSHEET:
        cellcount = 1
        loggerInfo.debug("%s Read Excel row %d " % (dbname, rowcount))

        # 遍历模版文件列信息
        for DBCELL in DBWROW:
            loggerInfo.debug("%s Read Excel row %d cell %d" % (dbname, rowcount, cellcount))

            # 设置Excel文件表格头内容
            if rowcount == 1 and cellcount == 1:
                loggerInfo.debug("Begin write sheet title info")

                dbcheckws.cell(column=cellcount, row=newrownum,
                               value='%s' % REPORTINFO.TITLE.get('REPORTTITLE')
                               .replace(DB_NAME_TEMPLATE, DB_NAME).replace(LASTDAY_TEMPLATE, LASTDAY))

                # 表格头内容特殊处理
                REPORTSTYLE.TITLEMERGE(dbcheckws, 1, 1, True)
                TitleCell = dbcheckws['A1']
                # TitleCell.border = TABLEBORDER()
                TitleCell.font = REPORTSTYLE.TITLEFONT('SHEETTITLE')
                TitleCell.fill = REPORTSTYLE.TITLEFILL('SHEETTITLE')
                TitleCell.alignment = REPORTSTYLE.TITLEALIGN('SHEETTITLE')
                break

            # 判断第一列是否为空列，空列保留空列信息用于表格分割
            if cellcount == 1 and DBCELL.value is None:
                loggerInfo.debug("Begin write sheet title info")

                # 从实例运行负载后开始插入表格数据
                if newrownum > 2:

                    # 表格模块标题
                    loggerInfo.debug("Excel sheet table title info %s" % tabsplit)
                    if REPORTINFO.TITLE.get(tabsplit) is not None:

                        # 记录表格开始行号和表格行数
                        tablestartrow = newrownum
                        tablerows = 0
                        # 判断处理的文件是否存在，不存在记录缺失文件到表格中，存在进行结果解析
                        if FILEEXIST(LASTDAY, DB_NAME, tabsplit):
                            # 处理文件内容

                            for filecontent in READSQLRESULT(LASTDAY, DB_NAME, tabsplit):
                                tablerows = tablerows + 1

                                cellcontent = filecontent.split('|')
                                lastcell = ''
                                for celllength in range(0, len(cellcontent)):

                                    cellvalue = cellcontent[celllength].strip()
                                    # 判断是否是数字类型，字符类型写入数字
                                    if cellvalue.isdigit():
                                        dbcheckws.cell(column=cellcount, row=newrownum).value = int(cellvalue)
                                    else:
                                        dbcheckws.cell(column=cellcount, row=newrownum).value = cellvalue

                                    contentcell = dbcheckws.cell(row=newrownum, column=cellcount)

                                    # 标记实例运行状态非OPEN和运行时间少于两天的数据库
                                    if (tabsplit == 'INSTANCE_STATUS' and ((cellcount == 6 and cellvalue != 'OPEN') or (
                                                    cellcount == 7 and int(cellvalue) < 2))):
                                        contentcell.fill = REPORTSTYLE.TITLEFILL('ALERTTITLE')

                                    # TABLESPACE_USAGE 表空间使用率判断 90
                                    # FILESYSTEM_USAGE文件系统使用率判断 85
                                    # ASM_USAGE ASM空间使用率判断 80
                                    if (tabsplit == 'TABLESPACE_USAGE' and cellcount == 7 and float(cellvalue) > 90) \
                                            or (tabsplit == 'FILESYSTEM_USAGE' and cellcount == 5 and \
                                                            int((cellvalue.split('%'))[0]) > 85) \
                                            or (tabsplit == 'ASM_USAGE' and cellcount == 8 and float(cellvalue) > 80):
                                        contentcell.fill = REPORTSTYLE.TITLEFILL('ALERTTITLE')

                                    # 标识状态为非ONLINE的集群资源
                                    if (tabsplit == 'CLUSTER_STATUS' and cellcount == 3 and cellvalue.find(
                                            'ONLINE') < 0):
                                        contentcell.fill = REPORTSTYLE.TITLEFILL('ALERTTITLE')

                                    contentcell = dbcheckws.cell(row=newrownum, column=cellcount)
                                    contentcell.border = REPORTSTYLE.TABLEBORDER('TABLECONTENT')

                                    cellcount = cellcount + 1

                                newrownum = newrownum + 1
                                cellcount = 1

                            # 生成实例运行负载图表
                            if tabsplit == 'INSTANCE_LOAD':
                                dbcheckws.add_chart(dbchart.DBLOAD_CHART(dbcheckws, DB_NAME, LASTDAY, tablerows),
                                                    'H' + str(tablestartrow))

                        else:
                            dbcheckws.cell(column=cellcount, row=newrownum,
                                           value='%s' % './' + LASTDAY + '/' + DB_NAME + '/' + tabsplit + '.txt' + '读取异常，请检查文件是否存在')

                        # 增加空白列，用户表格模块分割
                        newrownum = newrownum + 1
                        cellcount = 1
                        dbcheckws.cell(column=cellcount, row=newrownum, value='%s' % '')
                    else:
                        dbcheckws.cell(column=cellcount, row=newrownum, value='%s' % '')
                else:
                    dbcheckws.cell(column=cellcount, row=newrownum, value='%s' % '')

            # 非第一列空列，退出行信息处理
            elif DBCELL.value is None:
                break
            # 处理非表格头的行配置信息
            else:
                dbcheckws.cell(column=cellcount, row=newrownum, value='%s' % DBCELL.value)

                # 表格模块分割列
                if REPORTINFO.TITLE.get(DBCELL.value) is not None:
                    tabsplit = DBCELL.value
                    dbcheckws.cell(column=cellcount, row=newrownum, value='%s' % REPORTINFO.TITLE.get(DBCELL.value))
                    REPORTSTYLE.TITLEMERGE(dbcheckws, newrownum, cellcount, False)

                    TitleCell = dbcheckws['A' + str(newrownum)]
                    TitleCell.font = REPORTSTYLE.TITLEFONT('AREATITLE')
                    TitleCell.fill = REPORTSTYLE.TITLEFILL('AREATITLE')
                    TitleCell.alignment = REPORTSTYLE.TITLEALIGN('AREATITLE')

                # 表格模块内容标题
                else:
                    TitleCell = dbcheckws.cell(row=newrownum, column=cellcount)
                    TitleCell.border = REPORTSTYLE.TABLEBORDER()
                    TitleCell.font = REPORTSTYLE.TITLEFONT('TABLETITLE')
                    TitleCell.fill = REPORTSTYLE.TITLEFILL('TABLETITLE')
                    TitleCell.alignment = REPORTSTYLE.TITLEALIGN('TABLETITLE')

            cellcount = cellcount + 1

        rowcount = rowcount + 1
        newrownum = newrownum + 1

    #表格结尾内容
    if FILEEXIST(LASTDAY, DB_NAME, tabsplit):
        # 处理文件内容
        cellcount = 1

        for filecontent in READSQLRESULT(LASTDAY, DB_NAME, tabsplit):
            tablerows = tablerows + 1

            cellcontent = filecontent.split('|')
            for celllength in range(0, len(cellcontent)):

                cellvalue = cellcontent[celllength].strip()
                dbcheckws.cell(column=cellcount, row=newrownum).value = cellvalue
                contentcell = dbcheckws.cell(row=newrownum, column=cellcount)
                contentcell.border = REPORTSTYLE.TABLEBORDER('TABLECONTENT')

                cellcount = cellcount + 1

            newrownum = newrownum + 1
            cellcount = 1

    else:
        dbcheckws.cell(column=cellcount, row=newrownum,
                       value='%s' % './' + LASTDAY + '/' + DB_NAME + '/' + tabsplit + '.txt' + '读取异常，请检查文件是否存在')
